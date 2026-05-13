# SPDX-License-Identifier: AGPL-3.0-or-later
import csv
import json
from pathlib import Path

from app.converters.base import BaseConverter
from app.converters.registry import register


@register(("xlsx", "csv"))
class XlsxToCsvConverter(BaseConverter):
    def convert(self, input_path: Path, output_path: Path, **kwargs) -> Path:
        import openpyxl

        wb = openpyxl.load_workbook(str(input_path), read_only=True, data_only=True)
        ws = wb.active
        with output_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if v is None else str(v) for v in row])
        return output_path


@register(("csv", "xlsx"))
class CsvToXlsxConverter(BaseConverter):
    def convert(self, input_path: Path, output_path: Path, **kwargs) -> Path:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        with input_path.open(newline="", encoding="utf-8") as f:
            for row in csv.reader(f):
                ws.append(row)
        wb.save(str(output_path))
        return output_path


@register(("csv", "json"))
class CsvToJsonConverter(BaseConverter):
    def convert(self, input_path: Path, output_path: Path, **kwargs) -> Path:
        with input_path.open(newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        output_path.write_text(json.dumps(rows, indent=2, ensure_ascii=False), encoding="utf-8")
        return output_path


@register(("json", "csv"))
class JsonToCsvConverter(BaseConverter):
    def convert(self, input_path: Path, output_path: Path, **kwargs) -> Path:
        data = json.loads(input_path.read_text(encoding="utf-8"))
        if not isinstance(data, list) or not data:
            raise ValueError("JSON must be a non-empty array of objects for CSV conversion.")
        fieldnames = list(data[0].keys())
        with output_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)
        return output_path
